# ============================================================
# modules/tonnage.py — модуль расчёта тоннажа
# ============================================================
import os
import math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import NETWORK_ORDER, TEMPLATE_TOTAL_HEADER, get_style


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color, bold=False, size=12):
    return Font(color=hex_color, bold=bold, size=size, name="Arial")


def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _thin_border():
    thin = Side(border_style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _round_half_up(value) -> int:
    """Округление: если дробная часть >= 0.5 — в большую сторону, иначе в меньшую."""
    if value is None:
        return 0
    value = float(value)
    if value >= 0:
        return int(math.floor(value + 0.5))
    return int(-math.floor(-value + 0.5))


def _short_label(net_code: str, azs_header: str) -> str:
    """Заменяет 'АЗС' на краткое название сети перед номером, напр. АЗС123 -> БП-123."""
    text = str(azs_header).strip()
    if text.upper().startswith("АЗС"):
        rest = text[3:].strip()
        return f"{net_code} {rest}" if rest else net_code
    return f"{net_code} {text}"


def calculate_tonnage(root_folder: str, network_templates: dict):
    """
    network_templates: {network_code: [template_path, ...]}
    Создаёт файл Тоннаж_<folder_name>.xlsx в root_folder.
    """
    folder_name = os.path.basename(root_folder)
    out_path = os.path.join(root_folder, f"Тоннаж_{folder_name}.xlsx")

    network_data = {}
    for net_code, tpl_paths in network_templates.items():
        net_result = {}
        for tpl_path in tpl_paths:
            if not os.path.exists(tpl_path):
                continue
            try:
                wb = load_workbook(tpl_path, data_only=True)
                try:
                    ws = wb.active
                    _extract_weights(ws, net_result)
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass
            except Exception as e:
                print(f"  [!] Ошибка чтения {tpl_path}: {e}")
        network_data[net_code] = net_result

    if not network_data:
        print("  [!] Тоннаж: нет данных для расчёта.")
        return None

    ordered_nets = [n for n in NETWORK_ORDER if n in network_data]
    remaining = [n for n in network_data if n not in ordered_nets]
    ordered_nets.extend(remaining)

    # --- Списки АЗС и итоги по сетям (округление по правилу "больше половины — вверх") ---
    all_azs_per_net = {}
    network_totals = {}
    for net_code in ordered_nets:
        data = network_data.get(net_code, {})
        all_azs_per_net[net_code] = sorted(data.keys())
        network_totals[net_code] = _round_half_up(sum(data.values()))

    grand_total = sum(network_totals.values())
    max_rows = max((len(v) for v in all_azs_per_net.values()), default=0)

    # --- Ширина колонок и высота строк ---
    #
    # ВАЖНО: Excel переносит текст с wrap_text ПО ГРАНИЦАМ СЛОВ (по
    # пробелам), а не просто "когда закончилось место в строке по счёту
    # символов". Например, "СМ Бойтик" — это два слова: если после "СМ "
    # не остаётся места под "Бойтик" целиком, перенос произойдёт именно
    # после "СМ", независимо от того, сколько символов теоретически могло
    # бы поместиться. Поэтому вместо деления общей длины строки на
    # "символов на строку" ниже честно симулируется перенос по словам —
    # так же, как это делает сам Excel.
    #
    # Ширина колонки в Excel хранится в "единицах символа шрифта по
    # умолчанию" (обычно Calibri 11), а не символах реального шрифта
    # ячейки (у нас Arial 16, где символ заметно шире) — поэтому ширина
    # переводится в пиксели по стандартной формуле Excel (px = units*7+5)
    # и уже в пикселях сравнивается с шириной текста в реальном шрифте.
    PX_PER_UNIT = 7
    UNIT_PADDING_PX = 5
    CELL_PADDING_PX = 8      # внутренние отступы ячейки
    CHAR_PX_FACTOR = 0.78    # ширина символа Arial/кириллица (с запасом, чтобы не занижать)
    HEIGHT_SAFETY_PT = 4     # небольшой запас в points на неточность оценки ширины

    def _char_px(font_size: float) -> float:
        return font_size * CHAR_PX_FACTOR

    def _text_width_px(text: str, font_size: float) -> float:
        return len(str(text)) * _char_px(font_size)

    def _px_to_width_units(px: float) -> float:
        return max(1.0, (px - UNIT_PADDING_PX) / PX_PER_UNIT)

    def _avail_px(width_units: float) -> float:
        return max(1.0, width_units * PX_PER_UNIT + UNIT_PADDING_PX - CELL_PADDING_PX)

    def _wrap_word_count(text: str, avail_px: float, font_size: float) -> int:
        """Симулирует перенос текста ПО СЛОВАМ в доступной ширине avail_px
        и возвращает реальное число строк — так же, как это делает Excel
        с wrap_text=True (а не по грубому счёту символов на строку)."""
        words = str(text).split(" ")
        if not words:
            return 1
        lines = 1
        cur_width = 0.0
        space_px = _char_px(font_size) * 0.5
        for i, word in enumerate(words):
            word_px = _text_width_px(word, font_size)
            add_px = word_px if cur_width == 0 else space_px + word_px
            if cur_width + add_px <= avail_px or cur_width == 0:
                cur_width += add_px
            else:
                lines += 1
                cur_width = word_px
        return lines

    # Для расчёта ШИРИНЫ колонки берём ширину текста с небольшим запасом
    # (15%) — чтобы реальный рендер Arial в Excel (который может немного
    # отличаться от нашей оценки) комфортно помещался, не впритык.
    WIDTH_SAFETY_MULT = 1.15
    # Для расчёта ВЫСОТЫ строки (сколько строк займёт текст при переносе)
    # берём почти всю доступную ширину колонки (92%), а не всю целиком —
    # небольшой запас на случай неточности, но не настолько агрессивный,
    # чтобы обычные короткие названия ошибочно считались "переносящимися"
    # и раздували высоту строки без необходимости.
    HEIGHT_AVAIL_FACTOR = 0.92

    def _lines_needed(text: str, col_width_units: float, font_size: float) -> int:
        if not text:
            return 1
        avail_for_height = _avail_px(col_width_units) * HEIGHT_AVAIL_FACTOR
        return max(1, _wrap_word_count(str(text), avail_for_height, font_size))

    WEIGHT_COL_WIDTH = 9
    MIN_NAME_COL_WIDTH = 12
    MAX_NAME_COL_WIDTH = 28  # верхний предел, чтобы таблица не "расползалась" при печати

    title_font_size = 18
    header_font_size = 16
    data_font_size = 16
    # Высота одной строки текста в points при data_font_size, плюс
    # запасная "полу-строка" на случай, если реальный рендер Excel/шрифта
    # окажется чуть шире нашей оценки — лучше строка будет на пару
    # пунктов выше нужного, чем текст обрежется снизу.
    LINE_HEIGHT = round(data_font_size * 1.3)

    # --- Ширина колонки названия АЗС/магазина — под самую длинную метку
    # этой сети (а не под название сети в шапке — оно должно просто
    # переноситься на 2 строки, как уже происходит у "Партнер Нефть",
    # а не раздувать колонку под себя, если названия АЗС короткие) ---
    name_col_width = {}
    for net_code in ordered_nets:
        needed = MIN_NAME_COL_WIDTH
        for azs in all_azs_per_net[net_code]:
            label = _short_label(net_code, azs)
            label_px = _text_width_px(label, data_font_size) * WIDTH_SAFETY_MULT + CELL_PADDING_PX
            needed = max(needed, _px_to_width_units(label_px))
        name_col_width[net_code] = min(needed, MAX_NAME_COL_WIDTH)



    tmp_path = out_path + ".tmp"
    wb_out = Workbook()
    try:
        ws_out = wb_out.active
        ws_out.title = "Тоннаж"

        col_offset = 1
        col_map = {}
        for net_code in ordered_nets:
            col_map[net_code] = (col_offset, col_offset + 1)
            col_offset += 2
        last_col_idx = col_offset - 1
        last_col = get_column_letter(last_col_idx)

        # --- Заголовок (название папки) ---
        ws_out.merge_cells(f"A1:{last_col}1")
        ws_out["A1"] = folder_name
        ws_out["A1"].font = _font("000000", bold=True, size=title_font_size)
        ws_out["A1"].alignment = _align()
        ws_out.row_dimensions[1].height = 28

        # --- Заголовки сетей ---
        header_lines = 1
        for net_code in ordered_nets:
            style = get_style(net_code)
            c1, c2 = col_map[net_code]
            cl1, cl2 = get_column_letter(c1), get_column_letter(c2)
            ws_out.merge_cells(f"{cl1}2:{cl2}2")
            cell = ws_out[f"{cl1}2"]
            cell.value = style["full_name"]
            cell.fill = _fill(style["header_fill"])
            cell.font = _font(style["header_font"], bold=True, size=header_font_size)
            cell.alignment = _align()
            merged_width = name_col_width[net_code] + WEIGHT_COL_WIDTH
            header_lines = max(header_lines, _lines_needed(style["full_name"], merged_width, header_font_size))
        # высота шапки — под самое длинное название сети, которое теперь
        # переносится на нужное число строк (колонка сужена под реальные
        # данные АЗС, а не под название сети)
        HEADER_LINE_HEIGHT = round(header_font_size * 1.3)
        ws_out.row_dimensions[2].height = max(36, header_lines * HEADER_LINE_HEIGHT + HEIGHT_SAFETY_PT)

        # --- Данные по АЗС ---
        # Высота строки данных ОДНА НА ВСЕ строки (не пересчитывается для
        # каждой строки отдельно) — иначе из-за небольших различий в
        # оценке ширины шрифта соседние строки получали разную высоту,
        # хотя весь текст в них умещался в одну строку. Берём максимум
        # по ВСЕЙ таблице один раз — так высота гарантированно достаточна
        # для самой длинной метки и при этом одинакова для всех строк.
        global_max_lines = 1
        for net_code in ordered_nets:
            for azs in all_azs_per_net[net_code]:
                label = _short_label(net_code, azs)
                global_max_lines = max(global_max_lines, _lines_needed(label, name_col_width[net_code], data_font_size))
        data_row_height = max(20, global_max_lines * LINE_HEIGHT + HEIGHT_SAFETY_PT)

        for row_i in range(max_rows):
            r = row_i + 3
            for net_code in ordered_nets:
                azs_list = all_azs_per_net[net_code]
                c_azs, c_wt = col_map[net_code]
                style = get_style(net_code)
                if row_i < len(azs_list):
                    azs_header = azs_list[row_i]
                    label = _short_label(net_code, azs_header)
                    total_wt = _round_half_up(network_data[net_code].get(azs_header, 0))

                    cell_azs = ws_out.cell(row=r, column=c_azs, value=label)
                    cell_wt = ws_out.cell(row=r, column=c_wt, value=total_wt)

                    for cell in (cell_azs, cell_wt):
                        cell.fill = _fill(style["row_fill"])
                        cell.font = _font(style["row_font"], size=data_font_size)
                        cell.alignment = _align()
                    cell_wt.number_format = "0"
            ws_out.row_dimensions[r].height = data_row_height

        # --- Итог по каждой сети (формула — пересчитывается при ручных правках) ---
        total_row = max_rows + 3
        for net_code in ordered_nets:
            style = get_style(net_code)
            c_azs, c_wt = col_map[net_code]
            wt_col_letter = get_column_letter(c_wt)

            label_cell = ws_out.cell(row=total_row, column=c_azs, value="Итого")
            if max_rows > 0:
                total_formula = f"=SUM({wt_col_letter}3:{wt_col_letter}{max_rows + 2})"
            else:
                total_formula = 0
            value_cell = ws_out.cell(row=total_row, column=c_wt, value=total_formula)

            for cell in (label_cell, value_cell):
                cell.fill = _fill(style["col_header_fill"])
                cell.font = _font(style["col_header_font"], bold=True, size=data_font_size)
                cell.alignment = _align()
            value_cell.number_format = "0"
        ws_out.row_dimensions[total_row].height = 22

        # --- Общий итог по всем сетям ---
        # Объединение делится строго пополам по числу колонок (last_col_idx всегда чётный,
        # т.к. на каждую сеть приходится 2 колонки), поэтому пропорция надписи и значения
        # не меняется при добавлении/удалении сетей.
        grand_row = total_row + 1
        half = last_col_idx // 2

        # --- Часть 1: надпись "ИТОГО ВСЕГО, кг:" ---
        if half > 1:
            ws_out.merge_cells(f"A{grand_row}:{get_column_letter(half)}{grand_row}")
        label_cell = ws_out.cell(row=grand_row, column=1, value="ОБЩИЙ ВЕС:")
        label_cell.font = _font("FFFFFF", bold=True, size=header_font_size)
        label_cell.fill = _fill("404040")
        label_cell.alignment = _align(h="center")

        # --- Часть 2: сам общий тоннаж (формула — сумма итогов всех сетей) ---
        value_start_col = half + 1
        if last_col_idx - half > 1:
            ws_out.merge_cells(
                f"{get_column_letter(value_start_col)}{grand_row}:{last_col}{grand_row}"
            )

        total_refs = [
            f"{get_column_letter(col_map[net_code][1])}{total_row}"
            for net_code in ordered_nets
        ]
        grand_formula = "=" + "+".join(total_refs) if total_refs else 0

        gt_cell = ws_out.cell(row=grand_row, column=value_start_col, value=grand_formula)
        gt_cell.font = _font("FFFFFF", bold=True, size=header_font_size)
        gt_cell.fill = _fill("404040")
        gt_cell.alignment = _align()
        gt_cell.number_format = "0"
        ws_out.row_dimensions[grand_row].height = 24

        # --- Ширина колонок: у каждой сети своя ширина под названия АЗС/
        # магазинов, колонка веса везде узкая ---
        for net_code in ordered_nets:
            c_azs, c_wt = col_map[net_code]
            ws_out.column_dimensions[get_column_letter(c_azs)].width = name_col_width[net_code]
            ws_out.column_dimensions[get_column_letter(c_wt)].width = WEIGHT_COL_WIDTH

        # --- Границы таблицы ---
        border_obj = _thin_border()
        for row in ws_out.iter_rows(min_row=1, max_row=grand_row, min_col=1, max_col=last_col_idx):
            for cell in row:
                cell.border = border_obj

        # --- Печать: книжная ориентация, всё по ширине на одной странице ---
        ws_out.page_setup.orientation = "portrait"
        ws_out.page_setup.fitToWidth = 1
        ws_out.page_setup.fitToHeight = 0
        ws_out.page_setup.fitToPage = True
        ws_out.sheet_properties.pageSetUpPr.fitToPage = True
        ws_out.print_options.horizontalCentered = True

        wb_out.save(tmp_path)
    finally:
        try:
            wb_out.close()
        except Exception:
            pass

    if os.path.exists(out_path):
        os.remove(out_path)
    os.rename(tmp_path, out_path)

    print(f"  [✓] Тоннаж: {out_path}")
    return out_path


def _extract_weights(ws, result: dict):
    azs_cols = {}
    for col_idx in range(5, ws.max_column + 1):
        header = ws.cell(row=4, column=col_idx).value
        if header and str(header).strip() != TEMPLATE_TOTAL_HEADER:
            azs_cols[col_idx] = str(header)

    if not azs_cols:
        return

    max_row = ws.max_row
    for col_idx, azs_header in azs_cols.items():
        total = 0.0
        for r in range(5, max_row + 1):
            qty_val = ws.cell(row=r, column=col_idx).value
            wt_val = ws.cell(row=r, column=4).value
            try:
                qty = float(qty_val) if qty_val else 0
                wt = float(wt_val) if wt_val else 0
                total += qty * wt
            except (TypeError, ValueError):
                pass
        existing = result.get(azs_header, 0)
        result[azs_header] = existing + total