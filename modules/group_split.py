# ============================================================
# modules/group_split.py — разделение шаблона по группам товара
#
# ОТСЛЕЖИВАНИЕ ИЗМЕНЕНИЙ:
#   Отслеживается хэш (MD5) каждого файла-шаблона. Если содержимое
#   шаблона изменилось (пользователь поправил цифры вручную) — все
#   части, которые используют этот шаблон, помечаются на
#   перегенерацию. Хэши хранятся в .azs_state.json под ключом
#   "template_hashes".
#
#   Физическое удаление заявки пользователем здесь больше НЕ
#   обрабатывается — это отслеживается заранее, на уровне главного
#   меню (main.py), которое блокирует операции 2 и 3, пока заявки
#   не будут пересобраны через «Запустить все операции» / п.1.
# ============================================================
import os
import re
import copy
import shutil
import stat
import time
import hashlib

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    COL_GROUP, NETWORK_ORDER, TEMPLATE_TOTAL_HEADER, get_style,
    GROUP_ORDER, GROUP_ORDER_TAIL,
)
from modules.state import load_state, save_state


# ── надёжное удаление папки ───────────────────────────────────

def _safe_rmtree(path: str, retries: int = 6, delay: float = 0.5) -> bool:
    def _onerror(func, p, exc_info):
        try:
            os.chmod(p, stat.S_IWRITE)
            func(p)
        except Exception:
            pass

    for attempt in range(1, retries + 1):
        try:
            shutil.rmtree(path, onerror=_onerror)
            return True
        except PermissionError as e:
            if attempt == retries:
                print(f"  [!] Не удалось удалить «{path}» ({retries} попыток): {e}")
                print(f"      Закройте файлы в Проводнике и повторите.")
                return False
            time.sleep(delay * attempt)
    return False


# ── путь к папке части 1С ──────────────────────────────────────

def _part_target_dir(root_folder: str, p_int: int) -> str:
    return os.path.join(
        root_folder,
        "1С_Загрузка товаров" if p_int == 1 else f"1С_Загрузка товаров часть {p_int}"
    )


_PART_DIR_RE = re.compile(r"^1С_Загрузка товаров часть (\d+)$")


def _existing_part_dirs(root_folder: str) -> dict:
    """Возвращает {номер_части(str): путь} для всех папок частей 1С,
    реально лежащих на диске в root_folder."""
    result = {}
    if not os.path.isdir(root_folder):
        return result
    for entry in os.scandir(root_folder):
        if not entry.is_dir():
            continue
        if entry.name == "1С_Загрузка товаров":
            result["1"] = entry.path
            continue
        m = _PART_DIR_RE.match(entry.name)
        if m:
            result[m.group(1)] = entry.path
    return result


# ── сортировка групп ──────────────────────────────────────────

def _sort_groups(groups: list) -> list:
    """Сортирует группы согласно config.settings.GROUP_ORDER (основной
    порядок) и GROUP_ORDER_TAIL (группы, которые всегда идут последними).
    Группы, отсутствующие в обоих списках, вставляются по алфавиту сразу
    после известных групп основного порядка — перед хвостом."""
    order_index = {g.lower(): i for i, g in enumerate(GROUP_ORDER)}
    tail_index = {g.lower(): i for i, g in enumerate(GROUP_ORDER_TAIL)}

    tail = [g for g in groups if g.lower() in tail_index]
    body = [g for g in groups if g.lower() not in tail_index]

    def body_key(g):
        gl = g.lower()
        if gl in order_index:
            return (0, order_index[gl], "")
        return (1, 0, gl)  # неизвестные группы — по алфавиту, после известных

    body_sorted = sorted(body, key=body_key)
    tail_sorted = sorted(tail, key=lambda g: tail_index[g.lower()])

    return body_sorted + tail_sorted


# ── копирование стиля ячейки ──────────────────────────────────

def _copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)


# ── конвертация xlsx → xls через Excel (win32com) ────────────
#
# ОПТИМИЗАЦИЯ: раньше Excel.Application запускался и закрывался
# заново на КАЖДЫЙ файл (это секунды накладных расходов на файл).
# Теперь используется одна общая сессия Excel на весь прогон
# split_by_groups — она открывается один раз и закрывается в конце.

class _ExcelSession:
    """Держит одно запущенное приложение Excel для пакетной конвертации
    xlsx → xls. Драматически быстрее, чем поднимать Excel на каждый файл."""

    def __init__(self):
        self.excel = None
        self._available = None  # None = ещё не проверяли

    def _ensure(self) -> bool:
        if self._available is not None:
            return self._available
        try:
            import win32com.client
            import pythoncom
            pythoncom.CoInitialize()
            # DispatchEx (а не Dispatch!) — принудительно создаёт НОВЫЙ, изолированный
            # процесс Excel. Обычный Dispatch() может подцепиться к уже запущенному
            # Excel пользователя (через Running Object Table), и тогда автоматизация
            # начинает конфликтовать с тем, чем человек занят в интерфейсе в этот
            # момент — отсюда нестабильные "Вызов был отклонен" (RPC_E_CALL_REJECTED)
            # на некоторых машинах.
            self.excel = win32com.client.DispatchEx("Excel.Application")
            self.excel.Visible = False
            self.excel.DisplayAlerts = False
            self.excel.Interactive = False
            self.excel.ScreenUpdating = False
            self.excel.EnableEvents = False
            try:
                # Не у всех версий/сборок Excel есть это свойство — не критично.
                self.excel.AskToUpdateLinks = False
            except Exception:
                pass
            self._available = True
        except Exception as e:
            print(f"  [!] pywin32 не найден ({e}) — файлы будут сохранены в .xlsx")
            self.excel = None
            self._available = False
        return self._available

    # Коды COM-ошибок, при которых Excel просто временно занят
    # (например, ещё не отпустил предыдущий файл) — имеет смысл
    # подождать и повторить, а не сразу сдаваться в .xlsx.
    _TRANSIENT_HRESULTS = {
        -2147418111,  # RPC_E_CALL_REJECTED  — «Вызов отклонен вызываемым объектом»
        -2147417846,  # RPC_E_SERVERCALL_RETRYLATER
    }

    def convert(self, xlsx_path: str, xls_path: str, retries: int = 5, delay: float = 1.5) -> bool:
        if not self._ensure():
            return False
        import pythoncom
        import time

        last_err = None
        for attempt in range(1, retries + 1):
            try:
                wb = self.excel.Workbooks.Open(os.path.abspath(xlsx_path))
                try:
                    wb.SaveAs(os.path.abspath(xls_path), FileFormat=56)
                finally:
                    wb.Close(False)
                try:
                    os.remove(xlsx_path)
                except OSError:
                    pass
                return True
            except Exception as e:
                last_err = e
                hresult = getattr(e, "hresult", None) or (e.args[0] if e.args else None)
                if hresult in self._TRANSIENT_HRESULTS and attempt < retries:
                    # даём Excel «отдышаться»: прокачиваем очередь COM-сообщений
                    # и ждём перед повторной попыткой
                    try:
                        pythoncom.PumpWaitingMessages()
                    except Exception:
                        pass
                    time.sleep(delay)
                    continue
                break

        print(f"  [!] win32com: {last_err} — файл оставлен как .xlsx")
        return False

    def close(self):
        if self.excel is not None:
            try:
                self.excel.Quit()
            except Exception:
                pass
            self.excel = None
            try:
                import pythoncom
                pythoncom.CoUninitialize()
            except Exception:
                pass


# ── MD5-хэш файла (быстро, читаем блоками) ───────────────────

def _file_md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ── сбор данных из шаблона ────────────────────────────────────

def _read_template_data(ws) -> dict:
    azs_headers = []
    total_col_idx = None
    for col_idx in range(5, ws.max_column + 1):
        h = ws.cell(row=4, column=col_idx).value
        h_str = str(h).strip() if h else ""
        if h_str == TEMPLATE_TOTAL_HEADER:
            total_col_idx = col_idx
        elif h_str:
            azs_headers.append((col_idx, h_str))

    rows = []
    for r in range(5, ws.max_row + 1):
        grp_val = ws.cell(row=r, column=3).value
        grp = str(grp_val).strip() if grp_val else ""
        if not grp:
            continue
        cells = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cells.append((c, cell.value, cell))
        rows.append({"row_idx": r, "group": grp, "cells": cells})

    header_cells = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=4, column=c)
        header_cells.append((c, cell.value, cell))

    title_cell = ws.cell(row=1, column=1)

    return {
        "azs_headers": azs_headers,
        "total_col_idx": total_col_idx,
        "rows": rows,
        "header_cells": header_cells,
        "title_cell": title_cell,
    }


# ── запись xlsx (openpyxl) с полным копированием стилей ───────

def _write_group_xlsx(
        out_path: str,
        tpl_data: dict,
        group_rows: list,
        keep_azs_indices: list,
        style_cfg: dict,
        ws_src,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    azs_headers = tpl_data["azs_headers"]
    total_col_idx = tpl_data.get("total_col_idx")
    header_cells = tpl_data["header_cells"]
    title_cell = tpl_data["title_cell"]
    has_total = total_col_idx is not None
    total_cols = 4 + len(keep_azs_indices) + (1 if has_total else 0)

    # ── Строки 1-3: объединённый заголовок ───────────────────
    last_col_letter = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last_col_letter}3")
    dst_title = ws.cell(row=1, column=1)
    dst_title.value = title_cell.value
    _copy_cell_style(title_cell, dst_title)

    # ── Строка 4: шапка таблицы ───────────────────────────────
    for c_idx, val, src_cell in header_cells[:4]:
        dst = ws.cell(row=4, column=c_idx, value=val)
        _copy_cell_style(src_cell, dst)

    for out_col, ai in enumerate(keep_azs_indices, start=5):
        orig_col_idx, azs_name = azs_headers[ai]
        src_cell = ws_src.cell(row=4, column=orig_col_idx)
        dst = ws.cell(row=4, column=out_col, value=azs_name)
        _copy_cell_style(src_cell, dst)

    total_out_col = 5 + len(keep_azs_indices)  # первая колонка после последней АЗС
    if has_total:
        src_total_header = ws_src.cell(row=4, column=total_col_idx)
        dst = ws.cell(row=4, column=total_out_col, value=TEMPLATE_TOTAL_HEADER)
        _copy_cell_style(src_total_header, dst)

    # ── Строки данных ─────────────────────────────────────────
    first_azs_letter = "E"
    last_azs_letter = get_column_letter(4 + len(keep_azs_indices))
    for out_row, row_info in enumerate(group_rows, start=5):
        orig_cells = {c_idx: (val, src_cell) for c_idx, val, src_cell in row_info["cells"]}

        for c_idx in range(1, 5):
            val, src_cell = orig_cells.get(c_idx, (None, None))
            dst = ws.cell(row=out_row, column=c_idx, value=val)
            if src_cell:
                _copy_cell_style(src_cell, dst)

        for out_col, ai in enumerate(keep_azs_indices, start=5):
            orig_col_idx, _ = azs_headers[ai]
            val, src_cell = orig_cells.get(orig_col_idx, (None, None))
            dst = ws.cell(row=out_row, column=out_col, value=val)
            if src_cell:
                _copy_cell_style(src_cell, dst)

        if has_total:
            # Сумма считается только по колонкам АЗС, реально попавшим в
            # этот файл (а не по всем АЗС шаблона) — иначе цифра была бы
            # некорректной для конкретной части/группы.
            _, src_total_cell = orig_cells.get(total_col_idx, (None, None))
            formula = f"=SUM({first_azs_letter}{out_row}:{last_azs_letter}{out_row})"
            dst = ws.cell(row=out_row, column=total_out_col, value=formula)
            if src_total_cell:
                _copy_cell_style(src_total_cell, dst)
            dst.number_format = "0"

    # ── Ширина колонок ────────────────────────────────────────
    for c_idx in range(1, 5):
        letter = get_column_letter(c_idx)
        src_dim = ws_src.column_dimensions.get(letter)
        if src_dim and src_dim.width:
            ws.column_dimensions[letter].width = src_dim.width

    for out_col, ai in enumerate(keep_azs_indices, start=5):
        orig_col_idx, azs_name = azs_headers[ai]
        src_dim = ws_src.column_dimensions.get(get_column_letter(orig_col_idx))
        letter = get_column_letter(out_col)
        if src_dim and src_dim.width:
            ws.column_dimensions[letter].width = src_dim.width
        else:
            ws.column_dimensions[letter].width = max(len(azs_name) + 2, 10)

    if has_total:
        letter = get_column_letter(total_out_col)
        src_dim = ws_src.column_dimensions.get(get_column_letter(total_col_idx))
        if src_dim and src_dim.width:
            ws.column_dimensions[letter].width = src_dim.width
        else:
            ws.column_dimensions[letter].width = max(len(TEMPLATE_TOTAL_HEADER) + 2, 10)

    # ── Высота строк ──────────────────────────────────────────
    for r in range(1, 4):
        src_h = ws_src.row_dimensions[r].height
        ws.row_dimensions[r].height = src_h if src_h else 15

    src_h4 = ws_src.row_dimensions[4].height
    ws.row_dimensions[4].height = src_h4 if src_h4 else 30

    for out_row, row_info in enumerate(group_rows, start=5):
        orig_row_idx = row_info["row_idx"]
        src_h = ws_src.row_dimensions[orig_row_idx].height
        ws.row_dimensions[out_row].height = src_h if src_h else 19

    # ── Автофильтр ────────────────────────────────────────────
    ws.auto_filter.ref = f"A4:{last_col_letter}4"
    ws.freeze_panes = "A5"

    tmp = out_path + ".tmp"
    try:
        wb.save(tmp)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if os.path.exists(out_path):
        os.remove(out_path)
    os.rename(tmp, out_path)


def _collect_current_files(root_folder: str, networks: dict) -> dict:
    result = {}
    for folder_name, net_code in networks.items():
        net_folder = os.path.join(root_folder, folder_name)
        if not os.path.isdir(net_folder):
            continue
        files = []
        for f in os.listdir(net_folder):
            if f.lower().endswith((".xlsx", ".xlsm", ".xls")) and not f.lower().startswith("шаблон_"):
                files.append(f)
        if files:
            result[net_code] = sorted(files)
    return result


# ── фильтрация АЗС для конкретного набора файлов ───────────────

def _filter_azs_for_new_files(azs_headers: list, new_files: dict, net_code: str, suffix: str = None):
    if not new_files:
        return []

    new_for_net = new_files.get(net_code, [])
    if not new_for_net:
        return []

    if suffix and net_code == "МП":
        if suffix.endswith("_food"):
            new_for_net = [f for f in new_for_net
                           if os.path.splitext(f)[0].strip().lower().endswith("ф")]
        elif suffix.endswith("_nonfood"):
            new_for_net = [f for f in new_for_net
                           if os.path.splitext(f)[0].strip().lower().endswith("н")]
        if not new_for_net:
            return []

    new_keys = set()
    for fname in new_for_net:
        base = os.path.splitext(fname)[0].strip().lower()
        if net_code == "МП" and base and base[-1] in ("ф", "н"):
            base = base[:-1].strip()
        new_keys.add(base)

    matched_indices = []
    for ai, (col_idx, azs_name) in enumerate(azs_headers):
        azs_lower = azs_name.lower()
        azs_suffix = azs_lower.replace("азс", "").strip()
        for key in new_keys:
            if azs_lower == f"азс{key}" or azs_lower == f"азс {key}":
                matched_indices.append(ai)
                break
            if azs_suffix == key:
                matched_indices.append(ai)
                break
            azs_digits = "".join(c for c in azs_lower if c.isdigit())
            key_digits = "".join(c for c in key if c.isdigit())
            if azs_digits and key_digits and azs_digits == key_digits:
                matched_indices.append(ai)
                break

    return matched_indices


# ── проверка изменений хэшей шаблонов ────────────────────────

def _check_template_hashes(
    root_folder: str,
    template_paths: dict,
    state: dict,
) -> tuple[dict, set, set]:
    """
    Считает MD5 каждого шаблона и сравнивает с сохранённым.
    Возвращает:
        new_hashes    — актуальный словарь {path: md5} для сохранения
        changed_set   — set путей шаблонов, которые изменились
        vanished_set  — set путей шаблонов, которые раньше существовали
                         (были в сохранённых хэшах), а сейчас пропали —
                         например, файл "Шаблон_МП_nonfood.xlsx" удалён
                         template_builder'ом, т.к. заявок больше нет.
                         Такие пути не попадают ни в template_paths (там
                         только реально существующие файлы), ни в
                         changed_set — и поэтому раньше части, которые
                         использовали только этот (исчезнувший) шаблон,
                         никогда не помечались на пересборку: их .xls
                         оставались на диске «осиротевшими» навсегда.
    """
    saved_hashes: dict = state.get("template_hashes", {})
    new_hashes: dict = {}
    changed_set: set = set()

    seen_paths: set = set()
    for net_code, paths in template_paths.items():
        for path in paths:
            if not os.path.exists(path):
                continue
            seen_paths.add(path)
            md5 = _file_md5(path)
            new_hashes[path] = md5
            if saved_hashes.get(path) != md5:
                changed_set.add(path)

    vanished_set = {p for p in saved_hashes if p not in seen_paths and not os.path.exists(p)}

    return new_hashes, changed_set, vanished_set


# ── определение сети по пути к шаблону ─────────────────────────

def _net_code_from_template_path(path: str) -> str:
    """Извлекает код сети из имени файла шаблона,
    напр. 'Шаблон_МП_nonfood.xlsx' -> 'МП', 'Шаблон_БП.xlsx' -> 'БП'."""
    base = os.path.splitext(os.path.basename(path))[0]
    prefix = "Шаблон_"
    if base.startswith(prefix):
        base = base[len(prefix):]
    for suf in ("_food", "_nonfood"):
        if base.endswith(suf):
            return base[: -len(suf)]
    return base


# ── какие части затронуты изменёнными шаблонами ───────────────

def _parts_affected_by_template_change(
    parts_registry: dict,
    changed_paths: set,
    template_paths: dict,
) -> list:
    """
    Возвращает список номеров частей (str), которые нужно перегенерировать
    из-за изменения содержимого шаблона.

    Логика: часть содержит файлы заявок → файлы заявок принадлежат сети →
    у сети есть шаблон → если шаблон изменился, часть нужно перегенерировать.
    """
    # Строим обратную карту: net_code → set шаблонов
    net_to_templates: dict[str, set] = {}
    for net_code, paths in template_paths.items():
        net_to_templates[net_code] = set(paths)

    affected = []
    for p_num, p_files in parts_registry.items():
        for net_code in p_files:
            templates_for_net = net_to_templates.get(net_code, set())
            if templates_for_net & changed_paths:   # пересечение
                if p_num not in affected:
                    affected.append(p_num)
                break
    return affected


def _parts_affected_by_vanished_templates(
    parts_registry: dict,
    vanished_paths: set,
) -> list:
    """
    Возвращает список номеров частей, которые ссылаются на сеть исчезнувшего
    шаблона (файл шаблона удалён, т.к. заявок для этой категории больше нет).
    Сеть определяется по имени файла, а не по template_paths — у исчезнувшего
    шаблона там уже нет записи.
    """
    if not vanished_paths:
        return []
    vanished_nets = {_net_code_from_template_path(p) for p in vanished_paths}

    affected = []
    for p_num, p_files in parts_registry.items():
        if any(net_code in p_files for net_code in vanished_nets):
            affected.append(p_num)
    return affected







def split_by_groups(
        root_folder: str,
        template_paths: dict,
        networks: dict = None,
        part_suffix: str = None,
        _aux_callback=None,
        skip_check: bool = False,
):
    """
    Разделяет объединённые шаблоны по товарным группам 1С.

    Отслеживает два вида изменений:
      1. Новые файлы заявок на диске       → новая часть
      2. Изменение содержимого шаблона     → пересборка всех затронутых частей

    Физическое удаление заявок с диска сюда не доходит — оно
    проверяется заранее в главном меню (main.py), которое в этом
    случае блокирует вызов данной операции.
    """
    # 1. Сканируем заявочники на диске
    current_files = _collect_current_files(root_folder, networks) if networks else {}

    # 1б. Проверяем: для файлов не зарегистрированных ни в одной части
    #     смотрим есть ли их АЗС-колонка в шаблоне.
    #     Нет в parts + нет в шаблоне → операция 1 не запускалась → блокируем.
    #     Нет в parts + есть в шаблоне → операция 1 уже выполнена → пропускаем.
    #     При skip_check=True (запуск всех операций) проверка пропускается.
    if not skip_check:
        state_pre = load_state(root_folder)
        parts_pre = state_pre.get("parts", {})
        registered_pre: dict = {}
        for p_files in parts_pre.values():
            for net_code, f_list in p_files.items():
                registered_pre.setdefault(net_code, set()).update(f_list)

        # Собираем АЗС-колонки из шаблонов (один раз на сеть).
        # Для МП food/nonfood — РАЗНЫЕ шаблоны с разным набором АЗС-колонок,
        # поэтому ключ — (net_code, suffix), а не просто net_code. Иначе
        # колонка "29" в food-шаблоне "закрывала" проверку и для nonfood,
        # и новый файл "29н" не блокировался, даже если на диске был только "29ф".
        template_azs_digits: dict = {}  # {(net_code, suffix): set of digit-strings}
        for net_code, f_list in current_files.items():
            known_set = registered_pre.get(net_code, set())
            unregistered = [f for f in f_list if f not in known_set]
            if not unregistered:
                continue
            # Читаем шаблоны этой сети, по отдельности для каждого суффикса
            for path in template_paths.get(net_code, []):
                if not os.path.exists(path):
                    continue
                suffix = None
                if net_code == "МП":
                    bname = os.path.basename(path).lower()
                    if "_food" in bname:
                        suffix = "food"
                    elif "_nonfood" in bname:
                        suffix = "nonfood"
                digits_key = (net_code, suffix)
                digits = template_azs_digits.setdefault(digits_key, set())
                try:
                    wb = load_workbook(path, data_only=True, read_only=True)
                    try:
                        ws = wb.active
                        for col_idx in range(5, ws.max_column + 1):
                            h = ws.cell(row=4, column=col_idx).value
                            if h:
                                d = "".join(c for c in str(h) if c.isdigit())
                                if d:
                                    digits.add(d)
                    finally:
                        try:
                            wb.close()
                        except Exception:
                            pass
                except Exception:
                    pass

        not_in_template = []
        for net_code, f_list in current_files.items():
            known_set = registered_pre.get(net_code, set())
            for fname in f_list:
                if fname in known_set:
                    continue  # уже в parts
                base = os.path.splitext(fname)[0].strip().lower()
                file_suffix = None
                if net_code == "МП" and base and base[-1] in ("ф", "н"):
                    file_suffix = "food" if base[-1] == "ф" else "nonfood"
                    base = base[:-1].strip()
                digits = template_azs_digits.get((net_code, file_suffix))
                if digits is None:
                    # Шаблона для этой сети/категории вообще не существует —
                    # значит операция 1 для неё ни разу не запускалась (это
                    # НЕ «всё уже обработано», как считалось раньше). Пример:
                    # для АЗС уже есть файл «29ф» (шаблон food собран), а
                    # «29н» — первая nonfood-заявка этой АЗС: шаблона nonfood
                    # ещё нет, и без этой проверки файл не блокировался, а
                    # тихо создавал пустую «новую часть» без реальных данных.
                    not_in_template.append(f"  • [{net_code}] {fname}")
                    continue
                file_digits = "".join(c for c in base if c.isdigit())
                if file_digits and file_digits not in digits:
                    not_in_template.append(f"  • [{net_code}] {fname}")

        if not_in_template:
            print("\n[!] Есть новые заявки, которые ещё не обработаны:")
            for u in not_in_template:
                print(u)
            print("\n[!] Сначала запустите операцию 1 — «Обработать заявки и собрать шаблоны».")
            return None

    # 2. Читаем состояние
    state = load_state(root_folder)
    parts_registry = state.get("parts", {})

    # ── 2а. Проверяем хэши шаблонов ──────────────────────────
    new_hashes, changed_templates, vanished_templates = _check_template_hashes(
        root_folder, template_paths, state
    )
    if changed_templates:
        names = [os.path.basename(p) for p in changed_templates]
        print(f"  [~] Шаблоны изменены: {', '.join(names)} — пересборка затронутых частей")
    if vanished_templates:
        names = [os.path.basename(p) for p in vanished_templates]
        print(f"  [~] Шаблоны удалены (нет заявок): {', '.join(names)} — очистка затронутых частей")

    parts_to_delete = []
    parts_to_regenerate = []

    # ── 2в. Проверяем физическое наличие папок частей ────────
    for p_num in parts_registry:
        if p_num in parts_to_delete or p_num in parts_to_regenerate:
            continue
        if not os.path.isdir(_part_target_dir(root_folder, int(p_num))):
            print(f"  [i] Папка Части {p_num} отсутствует на диске — пересоздаю.")
            parts_to_regenerate.append(p_num)

    # ── 2г. Части, затронутые изменением шаблона ─────────────
    if changed_templates:
        for p_num in _parts_affected_by_template_change(
            parts_registry, changed_templates, template_paths
        ):
            if p_num not in parts_to_regenerate and p_num not in parts_to_delete:
                parts_to_regenerate.append(p_num)

    # ── 2г-2. Части, затронутые исчезнувшим шаблоном ─────────
    # (категория, например МП nonfood, осталась без заявок и без шаблона —
    # часть, хранившая только такие файлы, иначе никогда не пересоберётся
    # и её устаревшие .xls останутся на диске навсегда)
    if vanished_templates:
        for p_num in _parts_affected_by_vanished_templates(
            parts_registry, vanished_templates
        ):
            if p_num not in parts_to_regenerate and p_num not in parts_to_delete:
                parts_to_regenerate.append(p_num)

    # ── 2д. Консолидация: сливаем осколочные части в Часть 1 ────
    # Если в реестре есть части >1, папки которых физически отсутствуют
    # на диске, это артефакт прерванного первого запуска (шаблоны
    # создавались по одному и split_by_groups вызывался несколько раз).
    # Сливаем все такие части в Часть 1.
    if len(parts_registry) > 1:
        missing_extra = [
            p for p in list(parts_registry.keys())
            if p != "1" and not os.path.isdir(_part_target_dir(root_folder, int(p)))
        ]
        if missing_extra:
            merged: dict = dict(parts_registry.get("1", {}))
            for p_num in missing_extra:
                for net_code, f_list in parts_registry[p_num].items():
                    existing = merged.get(net_code, [])
                    merged[net_code] = sorted(set(existing) | set(f_list))
                del parts_registry[p_num]
            parts_registry["1"] = merged
            print(f"  [i] Объединены осколочные части {missing_extra} → Часть 1")
            # Удаляем слитые части из списков пересборки/удаления —
            # их ключей больше нет в parts_registry, обращение к ним вызовет KeyError
            for p in missing_extra:
                if p in parts_to_regenerate:
                    parts_to_regenerate.remove(p)
                if p in parts_to_delete:
                    parts_to_delete.remove(p)
            # После слияния пересоздаём Часть 1
            if "1" not in parts_to_regenerate and "1" not in parts_to_delete:
                parts_to_regenerate.append("1")

    # ── 2д-2. Удаляем папки частей, пропавших из реестра ─────
    # Если заявка была удалена с диска, prune_deleted_files (state.py,
    # вызывается операцией 1) уже вычистила её из parts_registry — и,
    # если это была последняя заявка в части, номер части целиком
    # исчезает из реестра. Раньше физическая папка такой части при
    # этом никогда не удалялась (parts_to_delete объявлялся, но не
    # заполнялся) — она оставалась на диске со старыми файлами
    # навсегда. Здесь сверяем реестр с тем, что реально лежит на
    # диске, и убираем папки частей, которых в реестре больше нет.
    for p_num, dir_path in _existing_part_dirs(root_folder).items():
        if p_num in parts_registry:
            continue
        if _safe_rmtree(dir_path):
            print(f"  [i] Часть {p_num} больше не актуальна (заявки удалены) — папка удалена.")

    # ── 2е. Новые файлы заявок → новая часть ─────────────────
    registered_files: dict = {}
    for p_num, p_files in parts_registry.items():
        for net_code, f_list in p_files.items():
            registered_files.setdefault(net_code, set()).update(f_list)

    new_files: dict = {}
    has_new_files = False
    for net_code, f_list in current_files.items():
        reg_set = registered_files.get(net_code, set())
        added = [f for f in f_list if f not in reg_set]
        if added:
            new_files[net_code] = added
            has_new_files = True

    new_part_num = None
    if has_new_files:
        if not parts_registry:
            # Первый запуск — всё в Часть 1
            new_part_num = 1
        else:
            existing_nums = [int(k) for k in parts_registry.keys() if k.isdigit()]
            new_part_num = max(existing_nums) + 1 if existing_nums else 1
        parts_registry[str(new_part_num)] = new_files

    # ── 3. Сохраняем обновлённое состояние ───────────────────
    state["parts"] = parts_registry
    existing_nums = [int(k) for k in parts_registry.keys() if k.isdigit()]
    state["1c_parts"] = max(existing_nums) if existing_nums else 0

    known_flat: dict = {}
    for p_num, p_files in parts_registry.items():
        for net_code, f_list in p_files.items():
            known_flat.setdefault(net_code, []).extend(f_list)
    for net_code in known_flat:
        known_flat[net_code] = sorted(list(set(known_flat[net_code])))
    state["known_files"] = known_flat

    # Сохраняем актуальные хэши шаблонов
    state["template_hashes"] = new_hashes

    save_state(root_folder, state)

    # ── 4. Итоговый список частей к записи ───────────────────
    parts_to_write = []
    if has_new_files:
        parts_to_write.append(str(new_part_num))
    for p_num in parts_to_regenerate:
        if p_num not in parts_to_write:
            parts_to_write.append(p_num)

    if not parts_to_write and not parts_to_delete:
        print("  [i] Изменений не обнаружено.")
        return os.path.join(root_folder, "1С_Загрузка товаров")

    parts_to_write.sort(key=lambda x: int(x))

    # ── 5. Настройка порядка сетей и суффиксов ───────────────
    ordered_nets = [n for n in NETWORK_ORDER if n in template_paths]
    remaining = [n for n in template_paths if n not in ordered_nets]
    ordered_nets.extend(remaining)

    ordered_suffixes = []
    for net_code in ordered_nets:
        for path in template_paths.get(net_code, []):
            suffix = net_code
            bname = os.path.basename(path).lower()
            if "_food" in bname:
                suffix = f"{net_code}_food"
            elif "_nonfood" in bname:
                suffix = f"{net_code}_nonfood"
            if suffix not in ordered_suffixes:
                ordered_suffixes.append(suffix)

    # ── 6. Читаем данные из шаблонов (один раз для всех частей)
    all_data: dict = {}
    all_groups: set = set()

    for net_code in ordered_nets:
        style = get_style(net_code)
        for path in template_paths.get(net_code, []):
            if not os.path.exists(path):
                continue
            suffix = net_code
            bname = os.path.basename(path).lower()
            if "_food" in bname:
                suffix = f"{net_code}_food"
            elif "_nonfood" in bname:
                suffix = f"{net_code}_nonfood"

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb.active
                tpl_data = _read_template_data(ws)
                azs_headers = tpl_data["azs_headers"]

                by_group: dict = {}
                for row_info in tpl_data["rows"]:
                    grp = row_info["group"]
                    all_groups.add(grp)
                    by_group.setdefault(grp, []).append(row_info)

                for grp, rows in by_group.items():
                    azs_sums = [0.0] * len(azs_headers)
                    for row_info in rows:
                        orig_cells = {c_idx: val for c_idx, val, _ in row_info["cells"]}
                        for ai, (col_idx, _) in enumerate(azs_headers):
                            try:
                                azs_sums[ai] += float(orig_cells.get(col_idx) or 0)
                            except (TypeError, ValueError):
                                pass

                    all_data.setdefault(grp, {})[suffix] = {
                        "tpl_data": tpl_data,
                        "rows": rows,
                        "style": style,
                        "ws_src_path": path,
                        "azs_sums": azs_sums,
                    }
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

    # ── 7. Только группы с реальными заказами ────────────────
    sorted_groups = _sort_groups(list(all_groups))
    active_groups = []
    for grp in sorted_groups:
        net_entries = all_data.get(grp, {})
        if any(
            any(s > 0 for s in entry["azs_sums"])
            for entry in net_entries.values()
        ):
            active_groups.append(grp)

    excel_session = _ExcelSession()

    # ── кэш открытых исходных книг шаблонов ───────────────────
    # ОПТИМИЗАЦИЯ: раньше каждый шаблон-источник (ws_src) открывался
    # заново load_workbook() на КАЖДУЮ группу (десятки раз на один и
    # тот же файл). Теперь каждый путь открывается максимум один раз
    # и переиспользуется для всех групп/частей, закрывается в конце.
    wb_src_cache: dict = {}

    def _get_ws_src(path: str):
        cached = wb_src_cache.get(path)
        if cached is None:
            wb_src = load_workbook(path, data_only=True)
            wb_src_cache[path] = wb_src
            return wb_src.active
        return cached.active

    # ── внутренняя функция записи файлов одной части ─────────

    def _write_part_group_files(target_dir: str, grp: str, gi: int, part_files: dict):
        net_entries = all_data.get(grp, {})
        grp_dir = os.path.join(target_dir, f"{gi}_{grp}")
        os.makedirs(grp_dir, exist_ok=True)

        file_written = False
        for fi, suffix in enumerate(ordered_suffixes, start=1):
            entry = net_entries.get(suffix)
            if entry is None:
                continue

            tpl_data  = entry["tpl_data"]
            rows      = entry["rows"]
            style     = entry["style"]
            azs_sums  = entry["azs_sums"]

            net_code = suffix.split("_")[0]
            allowed = _filter_azs_for_new_files(
                tpl_data["azs_headers"], part_files, net_code, suffix
            )
            keep_azs = [i for i in allowed if azs_sums[i] > 0] if allowed else []

            if not keep_azs:
                continue

            ws_src = _get_ws_src(entry["ws_src_path"])
            base_name = f"{fi}_{suffix}_{grp}"

            if win32_available:
                tmp_xlsx = os.path.join(grp_dir, f"{base_name}_tmp.xlsx")
                xls_path = os.path.join(grp_dir, f"{base_name}.xls")
                _write_group_xlsx(tmp_xlsx, tpl_data, rows, keep_azs, style, ws_src)
                ok = excel_session.convert(tmp_xlsx, xls_path)
                file_name = f"{base_name}.xls" if ok else f"{base_name}_tmp.xlsx"
            else:
                file_name = f"{base_name}.xlsx"
                _write_group_xlsx(
                    os.path.join(grp_dir, file_name),
                    tpl_data, rows, keep_azs, style, ws_src
                )

            azs_names = [tpl_data["azs_headers"][i][1] for i in keep_azs]
            print(f"  [+] {suffix} / «{grp}»: {len(rows)} поз., {len(keep_azs)} АЗС "
                  f"({', '.join(azs_names)}) → {file_name}")
            file_written = True

        if not file_written and os.path.isdir(grp_dir):
            try:
                os.rmdir(grp_dir)
            except OSError:
                pass

    win32_available = excel_session._ensure()

    # ── 8. Записываем только нужные части ────────────────────
    last_dir = None
    for p_num in parts_to_write:
        p_int = int(p_num)
        label = "базовые/обновлённые заявки" if p_int == 1 else "новые/изменённые заявки"
        print(f"\n  Формирование Части {p_int} ({label}):")

        target_dir = _part_target_dir(root_folder, p_int)
        if os.path.isdir(target_dir):
            _safe_rmtree(target_dir)
        os.makedirs(target_dir, exist_ok=True)

        part_files = parts_registry[p_num]
        for gi, grp in enumerate(active_groups, start=1):
            _write_part_group_files(target_dir, grp, gi, part_files)

        if os.path.isdir(target_dir) and not os.listdir(target_dir):
            try:
                os.rmdir(target_dir)
            except OSError:
                pass
        else:
            last_dir = target_dir

    # ── освобождаем ресурсы: Excel и кэш открытых шаблонов ────
    excel_session.close()
    for wb_src in wb_src_cache.values():
        try:
            wb_src.close()
        except Exception:
            pass

    if _aux_callback and networks:
        _aux_callback(root_folder, networks)

    return last_dir if last_dir else os.path.join(root_folder, "1С_Загрузка товаров")