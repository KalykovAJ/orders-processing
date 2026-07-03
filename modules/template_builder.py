# ============================================================
# modules/template_builder.py — создание и заполнение шаблона
#
# ЛОГИКА ИНКРЕМЕНТАЛЬНОГО ОБНОВЛЕНИЯ:
#   • Если шаблон НЕ существует — создаём с нуля (все заявки).
#   • Если шаблон УЖЕ существует — только добавляем новые колонки
#     АЗС и удаляем колонки удалённых заявок. Всё, что пользователь
#     добавил вручную — остаётся нетронутым.
# ============================================================
import os
import hashlib
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    COL_NAME, COL_GROUP, COL_WEIGHT, COL_TOTAL, TEMPLATE_TOTAL_HEADER, get_style
)
from modules.reference import get_network_items
from modules.network_detect import get_azs_files, parse_azs_number
from modules.state import get_azs_hashes, save_azs_hashes


def _file_md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ── вспомогательные стилевые функции (без изменений) ─────────

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str, bold=False, size=11):
    return Font(color=hex_color, bold=bold, size=size, name="Arial")


def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _thin_border():
    thin_side = Side(border_style="thin", color="000000")
    return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def _set_col_widths(ws):
    limits = {1: (6, 10), 2: (35, 110), 3: (15, 45), 4: (10, 18)}
    default_limits = (10, 16)
    for col in ws.columns:
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in col:
            if cell.row <= 3:
                continue
            if cell.value is not None:
                val_str = str(cell.value).strip()
                if col_idx == 2 and cell.alignment and cell.alignment.wrap_text:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal or "left",
                        vertical=cell.alignment.vertical or "center",
                        wrap_text=False
                    )
                for line in val_str.split('\n'):
                    max_len = max(max_len, len(line))
        mn, mx = limits.get(col_idx, default_limits)
        ws.column_dimensions[col_letter].width = min(max(max_len + 5, mn), mx)


def _set_row_heights(ws):
    for r in range(1, 4):
        ws.row_dimensions[r].height = 15
    ws.row_dimensions[4].height = 30
    for row in ws.iter_rows(min_row=5):
        ws.row_dimensions[row[0].row].height = 19


def _apply_borders_to_all(ws):
    border_obj = _thin_border()
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border_obj


def _apply_total_column(ws, style: dict, first_data_col: int = 5):
    """Пересоздаёт колонку ИТОГО сразу после последней колонки АЗС —
    построчную сумму количества по всем текущим колонкам АЗС (формула,
    поэтому пересчитывается автоматически и при ручных правках
    количества пользователем).

    Колонка полностью удаляется и создаётся заново при каждом вызове,
    чтобы её позиция и диапазон формулы всегда соответствовали текущему
    набору колонок АЗС (после добавления/удаления заявок он сдвигается).
    """
    last_col = ws.max_column
    if last_col < first_data_col:
        return  # нет ни одной колонки АЗС — колонку ИТОГО не создаём

    total_col = last_col + 1
    header_cell = ws.cell(row=4, column=total_col, value=TEMPLATE_TOTAL_HEADER)
    header_cell.fill = _fill(style["col_header_fill"])
    header_cell.font = _font(style["col_header_font"], bold=True)
    header_cell.alignment = _align()

    first_letter = get_column_letter(first_data_col)
    last_letter = get_column_letter(last_col)
    r_fill = _fill(style["row_fill"])
    for r in range(5, ws.max_row + 1):
        cell = ws.cell(
            row=r, column=total_col,
            value=f"=SUM({first_letter}{r}:{last_letter}{r})"
        )
        cell.fill = r_fill
        cell.font = _font(style["row_font"])
        cell.alignment = _align()
        cell.number_format = "0"


def _update_header_merge(ws, style: dict):
    # Собираем диапазоны merge из строк 1-3
    to_remove = [str(m) for m in ws.merged_cells if m.min_row <= 3]

    for rng in to_remove:
        # openpyxl требует, чтобы все ячейки объединения физически существовали
        # перед вызовом unmerge_cells. Принудительно создаём недостающие.
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                _ = ws.cell(row=r, column=c)  # создаёт ячейку если её нет
        ws.unmerge_cells(rng)

    last_col = get_column_letter(ws.max_column)
    ws.merge_cells(f"A1:{last_col}3")
    cell = ws["A1"]
    cell.value = style["full_name"]
    cell.fill = _fill(style["header_fill"])
    cell.font = _font(style["header_font"], bold=True, size=14)
    cell.alignment = _align(h="left")


# ── чтение данных из заявочника ───────────────────────────────

def _fill_azs_column(ws, items_df: pd.DataFrame, azs_file: str,
                     col_idx: int, style: dict):
    try:
        with pd.ExcelFile(azs_file) as xf:
            df_azs = xf.parse(xf.sheet_names[0], header=None)

        total_col_idx = None
        name_col_idx = None
        header_row = None
        for ri, row in df_azs.iterrows():
            for ci, val in enumerate(row):
                if str(val).strip() == COL_TOTAL:
                    total_col_idx = ci
                    header_row = ri
                if str(val).strip() == "Наименование":
                    name_col_idx = ci
            if header_row is not None:
                break

        if header_row is None or total_col_idx is None:
            return

        data_map = {}
        for ri in range(header_row + 1, len(df_azs)):
            row = df_azs.iloc[ri]
            name = str(row.iloc[name_col_idx]).strip() if name_col_idx is not None else ""
            qty = row.iloc[total_col_idx]
            if name and name not in ("nan", ""):
                try:
                    data_map[name] = float(qty) if pd.notna(qty) else 0
                except (TypeError, ValueError):
                    data_map[name] = 0

        r_fill = _fill(style["row_fill"])
        for ri, row in items_df.iterrows():
            r = ri + 5
            item_name = str(row.get(COL_NAME, "")).strip()
            qty = data_map.get(item_name, 0)
            cell = ws.cell(row=r, column=col_idx, value=qty if qty else None)
            cell.fill = r_fill
            cell.alignment = _align()

    except Exception as e:
        print(f"  [!] Ошибка заявочника {os.path.basename(azs_file)}: {e}")


# ── первичное создание шаблона (все заявки) ───────────────────

def _create_new_template(
    ws, network_code: str, style: dict,
    items_df: pd.DataFrame, filtered_azs: list,
) -> dict:
    """Записывает заголовок, справочные колонки и все колонки АЗС.
    Возвращает {col_header: col_idx} уже записанных АЗС."""
    _write_header(ws, network_code, style, items_df)
    existing_azs = {}
    for fpath, azs_num in filtered_azs:
        col_header = f"АЗС{azs_num}"
        if col_header not in existing_azs:
            next_col = ws.max_column + 1
            existing_azs[col_header] = next_col
            cell = ws.cell(row=4, column=next_col, value=col_header)
            cell.fill = _fill(style["col_header_fill"])
            cell.font = _font(style["col_header_font"], bold=True)
            cell.alignment = _align()
        _fill_azs_column(ws, items_df, fpath, existing_azs[col_header], style)
    _apply_total_column(ws, style)
    return existing_azs


# ── инкрементальное обновление существующего шаблона ──────────

def _update_existing_template(
    output_path: str,
    network_code: str,
    style: dict,
    items_df: pd.DataFrame,
    filtered_azs: list,
    changed_files: set = None,
) -> tuple[set, set, set]:
    """
    Открывает существующий шаблон и:
      - добавляет колонки для новых заявок;
      - удаляет колонки заявок, файлы которых исчезли из папки;
      - обновляет данные в колонках, чьё содержимое на диске изменилось,
        а имя файла осталось прежним (см. changed_files) — иначе такую
        замену не ловит diff по заголовкам колонок, и старые цифры
        оставались бы в шаблоне навсегда.
    Возвращает (added_set, removed_set, refreshed_set) — имена файлов.
    """
    changed_files = changed_files or set()
    current_headers = {f"АЗС{azs_num}" for _, azs_num in filtered_azs}
    fpath_by_header = {f"АЗС{azs_num}": fpath for fpath, azs_num in filtered_azs}

    wb = load_workbook(output_path)
    ws = wb.active

    # ── убираем старую колонку ИТОГО перед диффом колонок АЗС —
    # иначе она будет ошибочно принята за колонку исчезнувшей заявки.
    # Пересоздаётся заново в конце, уже за актуальным набором АЗС.
    for col_idx in range(ws.max_column, 4, -1):
        if str(ws.cell(row=4, column=col_idx).value or "").strip() == TEMPLATE_TOTAL_HEADER:
            ws.delete_cols(col_idx)
            break

    # ── считываем уже существующие колонки АЗС из строки 4 ──
    existing_azs: dict[str, int] = {}   # {col_header: col_idx}
    for col_idx in range(5, ws.max_column + 1):
        val = ws.cell(row=4, column=col_idx).value
        if val:
            existing_azs[str(val)] = col_idx

    # ── определяем diff ──────────────────────────────────────
    to_add = [h for h in current_headers if h not in existing_azs]
    to_remove = [h for h in existing_azs if h not in current_headers]

    added_files: set = set()
    removed_files: set = set()

    # ── удаляем колонки исчезнувших заявок ───────────────────
    if to_remove:
        # Удаляем столбцы справа налево, чтобы индексы не сдвигались
        cols_to_del = sorted(
            [existing_azs[h] for h in to_remove], reverse=True
        )
        for col_idx in cols_to_del:
            ws.delete_cols(col_idx)
            removed_files.add(existing_azs[
                next(h for h, ci in existing_azs.items() if ci == col_idx)
            ])
        # Пересчитываем existing_azs после удалений
        existing_azs = {}
        for col_idx in range(5, ws.max_column + 1):
            val = ws.cell(row=4, column=col_idx).value
            if val:
                existing_azs[str(val)] = col_idx
        # Собираем имена файлов для лога
        removed_files = {h.replace("АЗС", "") for h in to_remove}

    # ── обновляем колонки, чьё содержимое изменилось на диске ──
    # (тот же файл заменили новым с другими цифрами — заголовок
    # колонки не меняется, поэтому его не ловят to_add/to_remove)
    to_refresh = [
        h for h in current_headers
        if h in existing_azs and h not in to_add
        and os.path.basename(fpath_by_header[h]) in changed_files
    ]
    for col_header in to_refresh:
        fpath = fpath_by_header[col_header]
        col_idx = existing_azs[col_header]
        _fill_azs_column(ws, items_df, fpath, col_idx, style)
    refreshed_files = {os.path.basename(fpath_by_header[h]) for h in to_refresh}

    # ── добавляем новые колонки АЗС ──────────────────────────
    for col_header in to_add:
        fpath = fpath_by_header[col_header]
        next_col = ws.max_column + 1
        existing_azs[col_header] = next_col
        cell = ws.cell(row=4, column=next_col, value=col_header)
        cell.fill = _fill(style["col_header_fill"])
        cell.font = _font(style["col_header_font"], bold=True)
        cell.alignment = _align()
        _fill_azs_column(ws, items_df, fpath, next_col, style)
        added_files.add(os.path.basename(fpath))

    # ── пересоздаём колонку ИТОГО сразу после последней АЗС ──
    _apply_total_column(ws, style)

    # ── обновляем шапку и оформление ─────────────────────────
    _update_header_merge(ws, style)
    _apply_borders_to_all(ws)
    _set_col_widths(ws)
    _set_row_heights(ws)

    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A4:{last_col}4"

    tmp_path = output_path + ".tmp"
    try:
        wb.save(tmp_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    os.replace(tmp_path, output_path)
    return added_files, removed_files, refreshed_files


# ── публичный интерфейс ───────────────────────────────────────

def build_template(
    refs: dict,
    network_code: str,
    network_folder: str,
    output_path: str,
    category: str = None,
    processed_files: set = None,
    root_folder: str = None,
) -> set:
    """
    Создаёт шаблон (первый запуск) или инкрементально обновляет его
    (последующие запуски). Ручные правки пользователя не затрагиваются.

    root_folder — если передан, включает отслеживание изменения
    содержимого заявочников по MD5 (см. changed_files в
    _update_existing_template): позволяет заметить, что пользователь
    заменил файл заявки новым с тем же именем, но другими цифрами.
    Если не передан (None) — поведение как раньше, без этой проверки.
    """
    if processed_files is None:
        processed_files = set()

    items_df = get_network_items(refs, network_code, category)
    style = get_style(network_code)
    is_mp = (network_code == "МП")

    # ── Собираем список заявочников из папки ──────────────────
    azs_files = get_azs_files(network_folder)
    filtered_azs: list[tuple[str, str]] = []
    for fpath in azs_files:
        fname = os.path.basename(fpath)
        if fname.lower().startswith("шаблон_"):
            continue
        azs_num, cat = parse_azs_number(fpath, is_mp)
        if is_mp and category and cat != category:
            continue
        filtered_azs.append((fpath, azs_num))

    # ── Нет заявок → удаляем шаблон если был ─────────────────
    if not filtered_azs:
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
                print(f"  [-] Шаблон удалён (нет заявок): {os.path.basename(output_path)}")
            except OSError as e:
                print(f"  [!] Не удалось удалить шаблон: {e}")
        return processed_files

    # ── Хэши содержимого заявочников (см. docstring root_folder) ─
    changed_content_files: set = set()
    all_hashes: dict = {}
    if root_folder is not None:
        all_hashes = get_azs_hashes(root_folder)
        for fpath, _ in filtered_azs:
            hkey = f"{network_code}|{os.path.basename(fpath)}"
            try:
                h = _file_md5(fpath)
            except OSError:
                continue
            if hkey in all_hashes and all_hashes[hkey] != h:
                changed_content_files.add(os.path.basename(fpath))
            all_hashes[hkey] = h

    label = network_code if not category else f"{network_code} ({category})"

    # ══════════════════════════════════════════════════════════
    # ВЕТКА А: шаблон ещё не существует → создаём с нуля
    # ══════════════════════════════════════════════════════════
    if not os.path.exists(output_path):
        print(f"  [{label}] Шаблон не найден — создаём с нуля...")
        tmp_path = output_path + ".tmp"
        wb = Workbook()
        try:
            ws = wb.active
            ws.title = network_code
            _create_new_template(ws, network_code, style, items_df, filtered_azs)
            _update_header_merge(ws, style)
            _apply_borders_to_all(ws)
            _set_col_widths(ws)
            _set_row_heights(ws)
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A4:{last_col}4"
            wb.save(tmp_path)
        finally:
            try:
                wb.close()
            except Exception:
                pass
        os.replace(tmp_path, output_path)
        azs_names = ", ".join(azs_num for _, azs_num in filtered_azs)
        print(f"  [+] {label}: создан, {len(filtered_azs)} заявок ({azs_names})")
        for fpath, _ in filtered_azs:
            processed_files.add(os.path.basename(fpath))
        if root_folder is not None:
            save_azs_hashes(root_folder, all_hashes)
        return processed_files

    # ══════════════════════════════════════════════════════════
    # ВЕТКА Б: шаблон существует → только diff
    # ══════════════════════════════════════════════════════════
    added, removed, refreshed = _update_existing_template(
        output_path, network_code, style, items_df, filtered_azs,
        changed_files=changed_content_files,
    )

    if added or removed or refreshed:
        parts = []
        if added:
            parts.append(f"+{len(added)} новых: {', '.join(sorted(added))}")
        if removed:
            parts.append(f"-{len(removed)} удалено: {', '.join(sorted(removed))}")
        if refreshed:
            parts.append(f"~{len(refreshed)} обновлено (изменилось содержимое): {', '.join(sorted(refreshed))}")
        print(f"  [~] {label}: обновлён ({'; '.join(parts)})")
    else:
        print(f"  [=] {label}: изменений нет, шаблон не тронут")

    for fpath, _ in filtered_azs:
        processed_files.add(os.path.basename(fpath))
    if root_folder is not None:
        save_azs_hashes(root_folder, all_hashes)
    return processed_files


# ── запись шапки и справочных колонок ────────────────────────

def _write_header(ws, network_code: str, style: dict, items_df: pd.DataFrame):
    h_fill = _fill(style["header_fill"])
    r_fill = _fill(style["row_fill"])

    ws["A1"].value = style["full_name"]
    ws["A1"].fill = h_fill
    ws["A1"].font = _font(style["header_font"], bold=True, size=14)
    ws["A1"].alignment = _align(h="left")

    col_h_fill = _fill(style["col_header_fill"])
    for ci, h in enumerate(["№", COL_NAME, COL_GROUP, COL_WEIGHT], start=1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill = col_h_fill
        cell.font = _font(style["col_header_font"], bold=True)
        cell.alignment = _align()

    for ri, row in items_df.iterrows():
        r = ri + 5
        ws.cell(row=r, column=1, value=ri + 1).alignment = _align()
        ws.cell(row=r, column=2, value=row.get(COL_NAME, "")).alignment = _align(h="left")
        ws.cell(row=r, column=3, value=row.get(COL_GROUP, "")).alignment = _align()
        ws.cell(row=r, column=4, value=row.get(COL_WEIGHT, "")).alignment = _align()
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = r_fill
            ws.cell(row=r, column=c).font = _font(style["row_font"], bold=False)