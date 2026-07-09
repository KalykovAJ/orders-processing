# ============================================================
# modules/aux_files.py — недовоз и список (только заявочники АЗС)
# ============================================================
import os
import re
from openpyxl import Workbook, load_workbook

from config.settings import NETWORK_ORDER


def _is_template_file(filename: str, net_code: str) -> bool:
    base = os.path.splitext(filename)[0]
    return base.lower().startswith("шаблон_")


def _get_azs_sheet_names_by_net(root_folder: str, network_mapping: dict, selected_nets: set = None) -> dict:
    by_net = {}
    for folder_name, net_code in network_mapping.items():
        if selected_nets is not None and net_code not in selected_nets:
            continue
        net_folder = os.path.join(root_folder, folder_name)
        if not os.path.isdir(net_folder):
            continue

        names = set()
        for f in os.listdir(net_folder):
            if not f.lower().endswith((".xlsx", ".xlsm", ".xls")):
                continue
            if _is_template_file(f, net_code):
                continue

            base = os.path.splitext(f)[0].strip()

            if net_code == "МП":
                if base.lower().endswith(("ф", "н")):
                    base = base[:-1].strip()

            base_clean = re.sub(rf'^{re.escape(net_code)}\s*', '', base, flags=re.IGNORECASE).strip()
            sheet_name = f"{net_code}{base_clean}"[:31]
            names.add(sheet_name)

        if names:
            by_net[net_code] = sorted(names)

    return by_net


def _ordered_sheet_names(by_net: dict) -> list:
    result = []
    seen_nets = set()
    for net_code in NETWORK_ORDER:
        if net_code in by_net:
            result.extend(by_net[net_code])
            seen_nets.add(net_code)
    for net_code, names in by_net.items():
        if net_code not in seen_nets:
            result.extend(names)
    return result


SERVICE_NAMES = {"Sheet", "Лист1", "Sheet1"}


def _sync_workbook(out_path: str, ordered_names: list, by_net: dict, keep_vba: bool = False) -> tuple:
    target = set(ordered_names)
    added = []
    removed = []

    if os.path.exists(out_path):
        wb = load_workbook(out_path, keep_vba=keep_vba)
    else:
        wb = Workbook()

    tmp_path = out_path + ".tmp"
    try:
        for sname in list(wb.sheetnames):
            if sname in SERVICE_NAMES:
                del wb[sname]

        existing = set(wb.sheetnames)

        net_last_idx = {}
        for i, sname in enumerate(wb.sheetnames):
            for net_code, names in by_net.items():
                if sname in names or sname.startswith(net_code):
                    net_last_idx[net_code] = i

        for sname in ordered_names:
            if sname in existing:
                continue
            owner_net = None
            for net_code, names in by_net.items():
                if sname in names:
                    owner_net = net_code
                    break

            insert_after = net_last_idx.get(owner_net, len(wb.sheetnames) - 1) if owner_net else len(wb.sheetnames) - 1
            wb.create_sheet(title=sname, index=insert_after + 1)
            if owner_net:
                net_last_idx[owner_net] = wb.sheetnames.index(sname)
            added.append(sname)

        for sname in list(wb.sheetnames):
            if sname not in target and sname not in SERVICE_NAMES:
                del wb[sname]
                removed.append(sname)

        wb.save(tmp_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if os.path.exists(out_path):
        os.remove(out_path)
    os.rename(tmp_path, out_path)
    return added, removed


def _print_sync_result(label: str, out_path: str, added: list, removed: list):
    parts = []
    if added:
        parts.append(f"+{len(added)} листов: {', '.join(added)}")
    if removed:
        parts.append(f"-{len(removed)}: {', '.join(removed)}")
    detail = f" ({'; '.join(parts)})" if parts else ""
    print(f"  [✓] {label}{detail}")


def create_or_update_nedovoz(root_folder: str, network_mapping: dict, selected_nets: list = None):
    folder_name = os.path.basename(root_folder)
    out_path = os.path.join(root_folder, f"Недовоз_{folder_name}.xlsx")

    nets_filter = set(selected_nets) if selected_nets is not None else None
    by_net = _get_azs_sheet_names_by_net(root_folder, network_mapping, selected_nets=nets_filter)
    if not by_net:
        print("  [!] Недовоз: заявочники не найдены, файл не создан.")
        return

    ordered = _ordered_sheet_names(by_net)
    added, removed = _sync_workbook(out_path, ordered, by_net, keep_vba=False)
    _print_sync_result("Недовоз синхронизирован", out_path, added, removed)


def create_or_update_spisok(root_folder: str, network_mapping: dict, selected_nets: list = None):
    out_path = os.path.join(root_folder, "Список.xlsm")

    nets_filter = set(selected_nets) if selected_nets is not None else None
    by_net = _get_azs_sheet_names_by_net(root_folder, network_mapping, selected_nets=nets_filter)
    if not by_net:
        print("  [!] Список: заявочники не найдены, файл не создан.")
        return

    ordered = _ordered_sheet_names(by_net)
    added, removed = _sync_workbook(out_path, ordered, by_net, keep_vba=True)
    _print_sync_result("Список синхронизирован", out_path, added, removed)